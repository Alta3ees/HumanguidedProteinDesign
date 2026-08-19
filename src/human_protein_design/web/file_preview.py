"""Lightweight local previews for common protein-design evidence files."""

from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any


STRUCTURE_SUFFIXES = {".pdb", ".ent", ".cif", ".mmcif", ".pqr"}
FASTA_SUFFIXES = {".fasta", ".fa", ".faa", ".fas"}
TABLE_SUFFIXES = {".csv", ".tsv"}
IMAGE_SUFFIXES = {".png", ".jpg", ".jpeg", ".svg", ".tif", ".tiff"}
TEXT_SUFFIXES = {".txt", ".md", ".log"}
JSON_SUFFIXES = {".json"}
PDF_SUFFIXES = {".pdf"}
ROSETTA_SCORE_SUFFIXES = {".sc"}
SPREADSHEET_SUFFIXES = {".xlsx"}


def classify_file(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix in STRUCTURE_SUFFIXES:
        return "structure"
    if suffix in FASTA_SUFFIXES:
        return "fasta"
    if suffix in TABLE_SUFFIXES:
        return "table"
    if suffix in IMAGE_SUFFIXES:
        return "image"
    if suffix in PDF_SUFFIXES:
        return "pdf"
    if suffix in JSON_SUFFIXES:
        return "json"
    if suffix in ROSETTA_SCORE_SUFFIXES:
        return "rosetta_score"
    if suffix in SPREADSHEET_SUFFIXES:
        return "spreadsheet"
    if suffix in TEXT_SUFFIXES:
        return "text"
    return "generic"


def _read_text(path: Path, limit: int = 120_000) -> str:
    with path.open("r", encoding="utf-8", errors="replace") as handle:
        return handle.read(limit)


def _parse_fasta(path: Path) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    header: str | None = None
    chunks: list[str] = []

    def flush() -> None:
        nonlocal header, chunks
        if header is None and not chunks:
            return
        sequence = "".join(chunks).replace(" ", "").upper()
        records.append(
            {
                "header": header or f"sequence_{len(records) + 1}",
                "sequence": sequence,
                "length": len(sequence),
            }
        )
        chunks = []

    for raw in _read_text(path).splitlines():
        line = raw.strip()
        if not line:
            continue
        if line.startswith(">"):
            flush()
            header = line[1:].strip() or f"sequence_{len(records) + 1}"
        else:
            chunks.append(line)
    flush()
    return records[:100]


def _parse_delimited(path: Path, delimiter: str) -> dict[str, Any]:
    rows: list[list[str]] = []
    with path.open("r", encoding="utf-8", errors="replace", newline="") as handle:
        reader = csv.reader(handle, delimiter=delimiter)
        for index, row in enumerate(reader):
            rows.append(row)
            if index >= 100:
                break
    headers = rows[0] if rows else []
    return {"headers": headers, "rows": rows[1:] if rows else [], "truncated": len(rows) >= 101}


def _parse_rosetta_score(path: Path) -> dict[str, Any]:
    headers: list[str] = []
    rows: list[list[str]] = []
    for raw in _read_text(path).splitlines():
        line = raw.strip()
        if not line.startswith("SCORE:"):
            continue
        tokens = line.split()[1:]
        if not tokens:
            continue
        if not headers:
            headers = tokens
            continue
        # Rosetta occasionally repeats the header in concatenated score files.
        if tokens == headers:
            continue
        rows.append(tokens)
        if len(rows) >= 100:
            break
    return {"headers": headers, "rows": rows, "truncated": len(rows) >= 100}


def _parse_xlsx(path: Path) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise RuntimeError("XLSX preview requires openpyxl.") from error

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows: list[list[str]] = []
    for index, row in enumerate(sheet.iter_rows(values_only=True)):
        rows.append(["" if value is None else str(value) for value in row])
        if index >= 100:
            break
    headers = rows[0] if rows else []
    return {
        "sheet": sheet.title,
        "sheets": workbook.sheetnames,
        "headers": headers,
        "rows": rows[1:] if rows else [],
        "truncated": len(rows) >= 101,
    }


def preview_file(path: Path) -> dict[str, Any]:
    """Return a browser-friendly preview description for a local project file."""
    kind = classify_file(path)
    payload: dict[str, Any] = {
        "kind": kind,
        "name": path.name,
        "suffix": path.suffix.lower(),
        "size_bytes": path.stat().st_size,
    }
    if kind == "fasta":
        payload["records"] = _parse_fasta(path)
    elif kind == "table":
        payload.update(_parse_delimited(path, "\t" if path.suffix.lower() == ".tsv" else ","))
    elif kind == "spreadsheet":
        payload.update(_parse_xlsx(path))
    elif kind == "rosetta_score":
        payload.update(_parse_rosetta_score(path))
    elif kind == "json":
        try:
            payload["data"] = json.loads(_read_text(path))
        except json.JSONDecodeError as error:
            payload["error"] = f"Invalid JSON: {error}"
            payload["text"] = _read_text(path)
    elif kind == "text":
        payload["text"] = _read_text(path)
    return payload
